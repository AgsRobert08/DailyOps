from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q
import json

from .models import Inscrito, Asistencia
from .whatsapp import enviar_whatsapp, mensaje_registro, mensaje_asistencia


def get_inscritos_filtrados(request):
    qs     = Inscrito.objects.all().order_by('nombre')
    genero = request.GET.get('genero', '').strip()
    zona   = request.GET.get('zona', '').strip()
    q      = request.GET.get('q', '').strip()
    if genero: qs = qs.filter(genero=genero)
    if zona:   qs = qs.filter(zona=zona)
    if q:      qs = qs.filter(Q(nombre__icontains=q) | Q(telefono__icontains=q))
    return qs


def hora_local():
    return timezone.localtime(timezone.now())


# ── INSCRIPCIONES ─────────────────────────────────────

@login_required
def inscripciones_list(request):
    inscritos = get_inscritos_filtrados(request)
    return render(request, "cursos/inscripciones/list.html", {
        "inscritos":  inscritos,
        "total":      inscritos.count(),
        "zonas":      Inscrito.ZONAS,
        "zona_sel":   request.GET.get('zona', ''),
        "genero_sel": request.GET.get('genero', ''),
        "q":          request.GET.get('q', ''),
    })


@login_required
def inscripcion_create(request):
    if request.method == "POST":
        try:
            inscrito = Inscrito(
                nombre             = request.POST.get("nombre"),
                genero             = request.POST.get("genero"),
                #zona               = request.POST.get("zona") or None,
                #subzona            = request.POST.get("subzona") or None,
                #otra_denominacion  = bool(request.POST.get("otra_denominacion")),
                #denominacion       = request.POST.get("denominacion") or None,
                #iglesia            = request.POST.get("iglesia") or None,
                #pastor             = request.POST.get("pastor") or None,
                telefono           = request.POST.get("telefono"),
                correo_electronico = request.POST.get("correo_electronico") or None,
                #grado              = request.POST.get("grado"),
                #periodo            = request.POST.get("periodo") or None,
                #monto              = request.POST.get("monto") or 0,
            )
            inscrito.save()
            try:
                from cursos.services import enviar_correo_registro
                enviar_correo_registro(inscrito)
            except Exception as e:
                print("Correo no enviado:", e)
                messages.warning(request, "Registro creado, pero el correo no pudo enviarse.")
            # WhatsApp al registrar
            '''
            if inscrito.telefono:
                try:
                    enviar_whatsapp(inscrito.telefono, mensaje_registro(inscrito))
                except Exception as e:
                    print(f"WhatsApp no enviado: {e}")
            '''
            messages.success(request, f'"{inscrito.nombre}" registrado correctamente.')
            return redirect("cursos:inscripciones")
        except Exception as e:
            messages.error(request, f"Error al registrar: {e}")

    return render(request, "cursos/inscripciones/form.html", {
        "editing": False,
        "zonas":   Inscrito.ZONAS,
        "grados":  Inscrito.GRADOS,
        "generos": Inscrito.GENEROS,
    })


@login_required
def inscripcion_edit(request, pk):
    inscrito = get_object_or_404(Inscrito, pk=pk)
    if request.method == "POST":
        inscrito.nombre             = request.POST.get("nombre")
        inscrito.genero             = request.POST.get("genero")
        #inscrito.zona               = request.POST.get("zona") or None
        #inscrito.subzona            = request.POST.get("subzona") or None
        #inscrito.otra_denominacion  = bool(request.POST.get("otra_denominacion"))
        #inscrito.denominacion       = request.POST.get("denominacion") or None
        #inscrito.iglesia            = request.POST.get("iglesia") or None
        #inscrito.pastor             = request.POST.get("pastor") or None
        inscrito.telefono           = request.POST.get("telefono")
        inscrito.correo_electronico = request.POST.get("correo_electronico") or None
        #inscrito.grado              = request.POST.get("grado")
        #inscrito.periodo            = request.POST.get("periodo") or None
        #inscrito.monto              = request.POST.get("monto") or 0
        inscrito.save()
        messages.success(request, "Inscrito actualizado correctamente.")
        return redirect("cursos:inscripciones")

    return render(request, "cursos/inscripciones/form.html", {
        "editing":  True,
        "inscrito": inscrito,
        "zonas":    Inscrito.ZONAS,
        "grados":   Inscrito.GRADOS,
        "generos":  Inscrito.GENEROS,
    })


@login_required
def inscripcion_delete(request, pk):
    inscrito = get_object_or_404(Inscrito, pk=pk)
    if request.method == "POST":
        nombre = inscrito.nombre
        inscrito.delete()
        messages.success(request, f'"{nombre}" eliminado.')
    return redirect("cursos:inscripciones")


@login_required
def inscripcion_qr(request, pk):
    inscrito = get_object_or_404(Inscrito, pk=pk)
    if not inscrito.qr_image:
        inscrito.generar_qr()
        Inscrito.objects.filter(pk=pk).update(qr_image=inscrito.qr_image.name)
    return render(request, "cursos/inscripciones/qr.html", {"inscrito": inscrito})

'''
@login_required
def inscripciones_pdf(request):
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("Instala reportlab: pip install reportlab", status=500)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista_Inscritos.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    cs = ParagraphStyle('c', parent=styles['Normal'], fontSize=8, leading=10)
    data = [["Nombre","Zona","Subzona","Otra denom.","Denominacion","Telefono","Correo","Grado","Monto"]]
    for a in get_inscritos_filtrados(request):
        data.append([
            Paragraph(a.nombre or "", cs),
            Paragraph(a.get_zona_display() or "", cs),
            Paragraph(a.subzona or "", cs),
            Paragraph("Si" if a.otra_denominacion else "No", cs),
            Paragraph(a.denominacion or "", cs),
            Paragraph(a.telefono or "", cs),
            Paragraph(a.correo_electronico or "", cs),
            Paragraph(a.get_grado_display() or "", cs),
            Paragraph(f"${a.monto}" if a.monto else "", cs),
        ])
    tabla = Table(data, colWidths=[95,55,65,60,95,75,140,110,50], repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#303854')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.4,colors.grey),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('TOPPADDING',(0,0),(-1,-1),3),
        ('BOTTOMPADDING',(0,0),(-1,-1),3),
    ]))
    doc.build([tabla])
    return response
'''
@login_required
def inscripciones_pdf(request):
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("Instala reportlab: pip install reportlab", status=500)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista_Inscritos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4,
                            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=20)

    styles = getSampleStyleSheet()
    cs = ParagraphStyle('c', parent=styles['Normal'], fontSize=9, leading=12)

    data = [["Nombre", "Género", "Teléfono", "Correo electrónico"]]

    for a in get_inscritos_filtrados(request):
        data.append([
            Paragraph(a.nombre or "", cs),
            Paragraph(a.get_genero_display() or "", cs),
            Paragraph(a.telefono or "", cs),
            Paragraph(a.correo_electronico or "—", cs),
        ])

    # A4 portrait ancho útil ~535pt — repartido entre 4 columnas
    tabla = Table(data, colWidths=[180, 70, 100, 185], repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#303854')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))

    doc.build([tabla])
    return response


# ── ASISTENCIA ────────────────────────────────────────

@login_required
def asistencia_list(request):
    ahora  = hora_local()
    fecha  = request.GET.get("fecha", ahora.date().isoformat())
    genero = request.GET.get("genero", "")
    asistencias = Asistencia.objects.filter(fecha=fecha).select_related("inscrito")
    if genero:
        asistencias = asistencias.filter(inscrito__genero=genero)
    return render(request, "cursos/asistencia/list.html", {
        "asistencias":     asistencias,
        "fecha":           fecha,
        "genero_sel":      genero,
        "total_presentes": asistencias.count(),
        "total_inscritos": Inscrito.objects.count(),
    })


@login_required
def asistencia_scan(request):
    return render(request, "cursos/asistencia/scanner.html")


@login_required
def asistencia_registrar(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Metodo no permitido"}, status=405)

    try:
        body   = json.loads(request.body)
        codigo = body.get("codigo", "").strip().replace('\r','').replace('\n','').replace('\t','').replace("'", "-")
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON invalido"}, status=400)

    if not codigo:
        return JsonResponse({"ok": False, "error": "Codigo vacio"})

    try:
        inscrito = Inscrito.objects.get(codigo=codigo)
    except Exception:
        return JsonResponse({"ok": False, "error": "QR no reconocido"})

    ahora = hora_local()
    hoy   = ahora.date()

    existente = Asistencia.objects.filter(inscrito=inscrito, fecha=hoy).first()
    if existente:
        return JsonResponse({
            "ok":      True,
            "nuevo":   False,
            "nombre":  inscrito.nombre,
            "hora":    existente.hora.strftime("%H:%M"),
            "mensaje": "Ya registrado hoy",
        })

    asistencia = Asistencia.objects.create(
        inscrito = inscrito,
        fecha    = hoy,
        hora     = ahora.time(),
        asistio  = True,
    )

    return JsonResponse({
        "ok":    True,
        "nuevo": True,
        "nombre": inscrito.nombre,
        "zona":   inscrito.get_zona_display() or "-",
        "grado":  inscrito.get_grado_display(),
        "hora":   asistencia.hora.strftime("%H:%M"),
    })


@login_required
def asistencia_delete(request, pk):
    asistencia = get_object_or_404(Asistencia, pk=pk)
    if request.method == "POST":
        asistencia.delete()
        messages.success(request, "Asistencia eliminada.")
    return redirect("cursos:asistencia")

'''
@login_required
def asistencia_pdf(request):
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("Instala reportlab: pip install reportlab", status=500)

    ahora  = hora_local()
    fecha  = request.GET.get("fecha", ahora.date().isoformat())
    genero = request.GET.get("genero", "")
    asistencias = Asistencia.objects.filter(fecha=fecha).select_related("inscrito")
    if genero:
        asistencias = asistencias.filter(inscrito__genero=genero)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Asistencia_{fecha}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    cs = ParagraphStyle('c', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
    data = [["Nombre","Telefono","Zona","Subzona","Grado","Hora"]]
    for a in asistencias:
        data.append([
            Paragraph(a.inscrito.nombre or "", cs),
            Paragraph(a.inscrito.telefono or "", cs),
            Paragraph(a.inscrito.get_zona_display() or "", cs),
            Paragraph(a.inscrito.subzona or "", cs),
            Paragraph(a.inscrito.get_grado_display(), cs),
            Paragraph(a.hora.strftime("%H:%M"), cs),
        ])
    tabla = Table(data, colWidths=[150,85,90,80,150,60], repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#303854')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.4,colors.grey),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),4),
        ('BOTTOMPADDING',(0,0),(-1,-1),4),
    ]))
    doc.build([tabla])
    return response
'''

@login_required
def asistencia_pdf(request):
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("Instala reportlab: pip install reportlab", status=500)

    ahora  = hora_local()
    fecha  = request.GET.get("fecha", ahora.date().isoformat())
    genero = request.GET.get("genero", "")

    asistencias = Asistencia.objects.filter(fecha=fecha).select_related("inscrito")
    if genero:
        asistencias = asistencias.filter(inscrito__genero=genero)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Asistencia_{fecha}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4,
                            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=20)

    styles = getSampleStyleSheet()
    cs = ParagraphStyle('c', parent=styles['Normal'], fontSize=9, leading=12)

    data = [["#", "Nombre", "Género", "Teléfono", "Correo electrónico", "Hora"]]

    for i, a in enumerate(asistencias, start=1):
        data.append([
            Paragraph(str(i), cs),
            Paragraph(a.inscrito.nombre or "", cs),
            Paragraph(a.inscrito.get_genero_display() or "", cs),
            Paragraph(a.inscrito.telefono or "", cs),
            Paragraph(a.inscrito.correo_electronico or "—", cs),
            Paragraph(a.hora.strftime("%H:%M"), cs),
        ])

    # A4 portrait ancho útil ~535pt
    tabla = Table(data, colWidths=[25, 155, 60, 90, 155, 50], repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#303854')),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
        ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('ALIGN',          (0, 0), (0, -1),  'CENTER'),   # columna #
        ('ALIGN',          (5, 0), (5, -1),  'CENTER'),   # columna hora
        ('ALIGN',          (0, 0), (-1, 0),  'CENTER'),   # header
        ('ALIGN',          (1, 1), (-2, -1), 'LEFT'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',     (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 6),
        ('LEFTPADDING',    (0, 0), (-1, -1), 8),
    ]))

    doc.build([tabla])
    return response

@login_required
def buscar_inscrito(request):
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse([], safe=False)
    inscritos = Inscrito.objects.filter(
        Q(nombre__icontains=q) | Q(telefono__icontains=q)
    )[:10]
    return JsonResponse([{
        "id":      i.id,
        "nombre":  i.nombre,
        "telefono": i.telefono,
        "codigo":  str(i.codigo),
    } for i in inscritos], safe=False)


# ── WHATSAPP ──────────────────────────────────────────────

@login_required
def whatsapp_historial(request):
    from .models import MensajeWhatsApp
    mensajes = MensajeWhatsApp.objects.select_related('inscrito').all()

    # Filtros
    tipo   = request.GET.get('tipo', '')
    estado = request.GET.get('estado', '')
    q      = request.GET.get('q', '').strip()

    if tipo:
        mensajes = mensajes.filter(tipo=tipo)
    if estado:
        mensajes = mensajes.filter(estado=estado)
    if q:
        mensajes = mensajes.filter(
            Q(inscrito__nombre__icontains=q) | Q(numero__icontains=q)
        )

    return render(request, "cursos/whatsapp/historial.html", {
        "mensajes":    mensajes[:200],
        "total":       mensajes.count(),
        "total_ok":    mensajes.filter(estado='enviado').count(),
        "total_error": mensajes.filter(estado='error').count(),
        "tipo_sel":    tipo,
        "estado_sel":  estado,
        "q":           q,
    })


@login_required
def whatsapp_enviar(request):
    """Envío manual a un inscrito específico."""
    from .models import MensajeWhatsApp
    from .whatsapp import enviar_whatsapp

    if request.method == "POST":
        inscrito_id = request.POST.get("inscrito_id")
        mensaje_txt = request.POST.get("mensaje", "").strip()

        if not inscrito_id or not mensaje_txt:
            messages.error(request, "Selecciona un inscrito y escribe un mensaje.")
            return redirect("cursos:whatsapp_historial")

        inscrito = get_object_or_404(Inscrito, pk=inscrito_id)

        if not inscrito.telefono:
            messages.error(request, f"{inscrito.nombre} no tiene teléfono registrado.")
            return redirect("cursos:whatsapp_historial")

        resultado = enviar_whatsapp(
            numero   = inscrito.telefono,
            mensaje  = mensaje_txt,
            inscrito = inscrito,
            tipo     = 'manual',
        )

        if resultado["ok"]:
            messages.success(request, f"Mensaje enviado a {inscrito.nombre}.")
        else:
            messages.error(request, f"Error al enviar: {resultado.get('error')}")

        return redirect("cursos:whatsapp_historial")

    # GET — modal no necesita vista propia, redirige al historial
    return redirect("cursos:whatsapp_historial")

    # ── EXCEL ─────────────────────────────────────────────

@login_required
def inscripciones_excel(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Instala openpyxl: pip install openpyxl", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscritos"

    # Estilos
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill    = PatternFill("solid", start_color="303854", end_color="303854")
    header_align   = Alignment(horizontal="center", vertical="center")
    cell_font      = Font(name="Arial", size=10)
    cell_align_l   = Alignment(horizontal="left",   vertical="center")
    cell_align_c   = Alignment(horizontal="center", vertical="center")
    alt_fill       = PatternFill("solid", start_color="F8FAFC", end_color="F8FAFC")
    thin           = Side(style="thin", color="E2E8F0")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Nombre", "Género", "Teléfono", "Correo electrónico"]
    col_widths = [5, 35, 14, 18, 35]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20

    for i, inscrito in enumerate(get_inscritos_filtrados(request), start=1):
        row  = i + 1
        fill = alt_fill if i % 2 == 0 else None
        data = [i, inscrito.nombre, inscrito.get_genero_display(),
                inscrito.telefono, inscrito.correo_electronico or "—"]

        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font   = cell_font
            cell.border = border
            cell.alignment = cell_align_c if col in (1,) else cell_align_l
            if fill:
                cell.fill = fill

    # Fila de totales
    total_row = ws.max_row + 2
    cell = ws.cell(row=total_row, column=1, value=f"Total inscritos: {ws.max_row - 1}")
    cell.font = Font(name="Arial", bold=True, size=10)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Lista_Inscritos.xlsx"'
    wb.save(response)
    return response


@login_required
def asistencia_excel(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Instala openpyxl: pip install openpyxl", status=500)

    ahora  = hora_local()
    fecha  = request.GET.get("fecha", ahora.date().isoformat())
    genero = request.GET.get("genero", "")

    asistencias = Asistencia.objects.filter(fecha=fecha).select_related("inscrito")
    if genero:
        asistencias = asistencias.filter(inscrito__genero=genero)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Asistencia {fecha}"

    # Estilos (idénticos al de inscritos)
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="303854", end_color="303854")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font    = Font(name="Arial", size=10)
    cell_align_l = Alignment(horizontal="left",   vertical="center")
    cell_align_c = Alignment(horizontal="center", vertical="center")
    alt_fill     = PatternFill("solid", start_color="F8FAFC", end_color="F8FAFC")
    thin         = Side(style="thin", color="E2E8F0")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["#", "Nombre", "Género", "Teléfono", "Correo electrónico", "Hora"]
    col_widths = [5, 35, 14, 18, 35, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20

    for i, a in enumerate(asistencias, start=1):
        row  = i + 1
        fill = alt_fill if i % 2 == 0 else None
        data = [
            i,
            a.inscrito.nombre,
            a.inscrito.get_genero_display(),
            a.inscrito.telefono,
            a.inscrito.correo_electronico or "—",
            a.hora.strftime("%H:%M"),
        ]

        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = cell_font
            cell.border    = border
            cell.alignment = cell_align_c if col in (1, 6) else cell_align_l
            if fill:
                cell.fill = fill

    total_row = ws.max_row + 2
    cell = ws.cell(row=total_row, column=1, value=f"Total presentes: {ws.max_row - 1}")
    cell.font = Font(name="Arial", bold=True, size=10)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Asistencia_{fecha}.xlsx"'
    wb.save(response)
    return response